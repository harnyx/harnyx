"""Minimal isolated source-extraction process; communicate only over stdin/stdout."""

from __future__ import annotations

import gzip
import io
import json
import os
import re
import resource
import sys
import zipfile
import zlib

MAX_ADDRESS_SPACE_BYTES = 512 * 1024 * 1024
MAX_CPU_SECONDS = 45
MAX_PDF_PAGES = 2_000
MAX_EXTRACTED_CHARACTERS = 5_000_000
MAX_XLSX_ZIP_ENTRIES = 10_000
MAX_XLSX_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 100


class ExtractionRejectedError(RuntimeError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


def main() -> None:
    try:
        _set_resource_limits()
        media_type, encoding, url = sys.argv[1:4]
        content = extract_content(sys.stdin.buffer.read(), media_type, encoding, url)
        if not content.strip():
            raise ExtractionRejectedError("source_unavailable", "source extraction returned no text")
        if len(content) > MAX_EXTRACTED_CHARACTERS:
            raise ExtractionRejectedError("source_extraction_limit", "source exceeds extracted-character limit")
        _write(b"O" + content.encode("utf-8"))
    except ExtractionRejectedError as exc:
        _write_error(exc.code, str(exc))
    except BaseException as exc:
        _write_error("source_unavailable", f"source extraction failed: {type(exc).__name__}: {exc}")


def extract_content(body: bytes, media_type: str, encoding: str, url: str) -> str:
    if encoding == "gzip":
        body = gzip.decompress(body)
    elif encoding == "deflate":
        body = zlib.decompress(body)
    normalized_url = url.casefold().split("?", 1)[0]
    if "spreadsheetml" in media_type or normalized_url.endswith(".xlsx"):
        return _extract_xlsx(body)
    if "macroenabled" in media_type or normalized_url.endswith(".xlsm"):
        raise ExtractionRejectedError("source_unavailable", "macro-enabled workbooks are unsupported")
    if "pdf" in media_type or normalized_url.endswith(".pdf"):
        import fitz

        document = fitz.open(stream=body, filetype="pdf")
        try:
            if document.page_count > MAX_PDF_PAGES:
                raise ExtractionRejectedError("source_extraction_limit", "PDF exceeds 2,000 pages")
            pieces: list[str] = []
            total = 0
            for index in range(1, document.page_count + 1):
                text = document.load_page(index - 1).get_text("text", sort=True)
                piece = f"[PDF page {index}]\n{text}"
                total += len(piece)
                if total > MAX_EXTRACTED_CHARACTERS:
                    raise ExtractionRejectedError("source_extraction_limit", "PDF exceeds extracted-character limit")
                pieces.append(piece)
            return "\n\n".join(pieces)
        finally:
            document.close()
    decoded = body.decode("utf-8", errors="replace")
    if "json" in media_type:
        try:
            return json.dumps(json.loads(decoded), ensure_ascii=False, indent=2)
        except json.JSONDecodeError as exc:
            raise ExtractionRejectedError("source_unavailable", "declared JSON source is malformed") from exc
    if "html" not in media_type and "<html" not in decoded[:1000].casefold():
        return decoded
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(decoded, "html.parser")
    for element in soup(["script", "style", "noscript", "template"]):
        element.decompose()
    lines: list[str] = []
    for table in soup.find_all("table"):
        for row_index, row in enumerate(table.find_all("tr")):
            cells = [re.sub(r"\s+", " ", cell.get_text(" ", strip=True)) for cell in row.find_all(["th", "td"])]
            if cells:
                prefix = "HEADER" if row_index == 0 or row.find("th") is not None else "ROW"
                lines.append(prefix + "\t" + "\t".join(cells))
        table.decompose()
    lines.extend(line.strip() for line in soup.get_text("\n").splitlines() if line.strip())
    return "\n".join(lines)


def _extract_xlsx(body: bytes) -> str:
    _preflight_xlsx_zip(body)
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True, keep_links=False)
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        raise ExtractionRejectedError("source_unavailable", "invalid or unsupported XLSX workbook") from exc
    try:
        lines: list[str] = []
        total = 0
        for worksheet in workbook.worksheets:
            heading = (
                "XLSX_WORKSHEET\t"
                f"title={json.dumps(worksheet.title, ensure_ascii=False)}\tstate={worksheet.sheet_state}"
            )
            lines.append(heading)
            total += len(heading) + 1
            for row in worksheet.iter_rows():
                populated = [
                    f"{cell.coordinate}={_xlsx_cell_text(cell.value)}" for cell in row if cell.value is not None
                ]
                if not populated:
                    continue
                line = f"XLSX_ROW\tworksheet={json.dumps(worksheet.title, ensure_ascii=False)}\trow={row[0].row}\t"
                line += "\t".join(populated)
                total += len(line) + 1
                if total > MAX_EXTRACTED_CHARACTERS:
                    raise ExtractionRejectedError(
                        "source_extraction_limit",
                        "XLSX exceeds extracted-character limit",
                    )
                lines.append(line)
        return "\n".join(lines)
    except ExtractionRejectedError:
        raise
    except Exception as exc:
        raise ExtractionRejectedError("source_unavailable", "XLSX row extraction failed") from exc
    finally:
        workbook.close()


def _preflight_xlsx_zip(body: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(body)) as archive:
            entries = archive.infolist()
    except zipfile.BadZipFile as exc:
        raise ExtractionRejectedError("source_unavailable", "XLSX is not a valid Open XML archive") from exc
    if len(entries) > MAX_XLSX_ZIP_ENTRIES:
        raise ExtractionRejectedError("source_extraction_limit", "XLSX archive contains too many entries")
    if any(entry.flag_bits & 0x1 for entry in entries):
        raise ExtractionRejectedError("source_unavailable", "encrypted XLSX workbooks are unsupported")
    if any(entry.filename.casefold().endswith(".bin") for entry in entries):
        raise ExtractionRejectedError("source_unavailable", "macro-enabled XLSX workbooks are unsupported")
    if sum(entry.file_size for entry in entries) > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise ExtractionRejectedError("source_extraction_limit", "XLSX archive expands beyond 128 MiB")
    for entry in entries:
        if not entry.file_size:
            continue
        if not entry.compress_size or entry.file_size / entry.compress_size > MAX_XLSX_COMPRESSION_RATIO:
            raise ExtractionRejectedError("source_extraction_limit", "XLSX archive entry compression ratio is unsafe")


def _xlsx_cell_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def _write(payload: bytes) -> None:
    offset = 0
    while offset < len(payload):
        offset += os.write(1, payload[offset:])


def _write_error(code: str, message: str) -> None:
    try:
        payload = json.dumps({"code": code[:64], "message": message[:2_000]}).encode()
    except BaseException:
        payload = b'{"code":"source_unavailable","message":"source extraction failed"}'
    _write(b"E" + payload)


def _set_resource_limits() -> None:
    resource.setrlimit(resource.RLIMIT_CPU, (MAX_CPU_SECONDS, MAX_CPU_SECONDS))
    resource.setrlimit(resource.RLIMIT_AS, (MAX_ADDRESS_SPACE_BYTES, MAX_ADDRESS_SPACE_BYTES))


if __name__ == "__main__":
    main()
